"""
Build Lumina Streaming Co. headcount synthetic dataset.
8 sheets, 6 seeded findings for agents to surface.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

wb = Workbook()

# ── Styles ───────────────────────────────────────────────────────────────────
ARIAL       = Font(name='Arial', size=10)
BOLD        = Font(name='Arial', size=10, bold=True)
HDR_FONT    = Font(name='Arial', size=10, bold=True, color='FFFFFF')
TITLE_FONT  = Font(name='Arial', size=14, bold=True)
FLAG_FONT   = Font(name='Arial', size=10, bold=True, color='9C0006')

HDR_FILL    = PatternFill('solid', start_color='1F4E78')
SUB_FILL    = PatternFill('solid', start_color='DDEBF7')
FLAG_FILL   = PatternFill('solid', start_color='FFCCCC')
WARN_FILL   = PatternFill('solid', start_color='FFE699')
GOOD_FILL   = PatternFill('solid', start_color='C6EFCE')
TOTAL_FILL  = PatternFill('solid', start_color='F2F2F2')

THIN = Side(border_style='thin', color='B4B4B4')
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal='center', vertical='center')
LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
RIGHT  = Alignment(horizontal='right',  vertical='center')

CURRENCY = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
PCT      = '0.0%;(0.0%);"-"'
DATE_FMT = 'yyyy-mm-dd'
INT_FMT  = '#,##0'

def hdr(ws, row, col, val, w=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = BOX
    if w: ws.column_dimensions[get_column_letter(col)].width = w

def cell(ws, row, col, val, fmt=None, fill=None, bold=False, wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font = BOLD if bold else ARIAL
    c.border = BOX
    c.alignment = LEFT if wrap else CENTER if fmt == DATE_FMT else RIGHT if fmt in (CURRENCY, PCT, INT_FMT) else LEFT
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    return c

def total_row(ws, r, label, cols_sum, ncols):
    ws.cell(row=r, column=1, value=label).font = BOLD
    ws.cell(row=r, column=1).fill = TOTAL_FILL
    ws.cell(row=r, column=1).border = BOX
    for c in cols_sum:
        col_ltr = get_column_letter(c)
        f = ws.cell(row=r, column=c, value=f'=SUM({col_ltr}2:{col_ltr}{r-1})')
        f.font = BOLD; f.fill = TOTAL_FILL; f.number_format = CURRENCY; f.border = BOX

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1: README
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = 'README'
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 85

ws['A1'] = 'Lumina Streaming Co. — Headcount & People Cost Dataset'
ws['A1'].font = TITLE_FONT
ws.merge_cells('A1:B1')

rows = [
    ('', ''),
    ('Purpose', 'Synthetic dataset for developing the headcount agent system. All values are fabricated.'),
    ('Company', 'Lumina Streaming Co. — ~600 employees globally, ~400 in LuminaUS.'),
    ('Period', 'November 2026 (snapshot date November 30, 2026)'),
    ('Currency', 'USD. EMEA and APAC salaries converted at period-average FX rate.'),
    ('Scope', 'LuminaUS employee detail. EMEA and APAC at summary level in budget sheet.'),
    ('', ''),
    ('Sheets', ''),
    ('  Headcount_Roster', 'Current (Nov-26) LuminaUS employees — 80 records across 5 departments'),
    ('  Prior_Roster', 'Oct-26 snapshot — includes 3 employees who departed in November'),
    ('  Headcount_Budget', 'Monthly headcount count and cost budget by department, full year FY2026'),
    ('  Open_Pipeline', 'Open requisitions as of Nov 30, 2026 — 15 open roles'),
    ('  Comp_Bands', 'Salary bands by department and level (min / midpoint / max)'),
    ('  Market_Benchmarks', 'External market comp benchmarks by role category (P25 / P50 / P75)'),
    ('  Attrition_Log', 'Employee departures in the trailing 3 months'),
    ('', ''),
    ('Seeded findings', 'Intentional issues for agents to surface:'),
    ('  1', 'Engineering headcount 5 over budget (30 actual vs 25 budgeted for LuminaUS sample)'),
    ('  2', 'Three open Engineering roles with target start dates > 30 days past — budget exposure $720K'),
    ('  3', 'Two Marketing managers paid > 15% above comp band midpoint (out-of-band high)'),
    ('  4', 'One Finance analyst paid 18% below comp band midpoint (out-of-band low)'),
    ('  5', 'Content team: 3 departures in November — annualised attrition 36%, above 2%/month flag'),
    ('  6', 'Sales: 1 open AE role 45 days past target start, commission plan not yet assigned'),
    ('', ''),
    ('Privacy rules', 'Never surface individual salaries in shared outputs. Aggregate to department level.'),
    ('Approval rule', 'All proposed comp adjustments require HR and Controller approval before actioning.'),
]
for i, (a, b) in enumerate(rows, start=2):
    ws[f'A{i}'] = a
    ws[f'B{i}'] = b
    ws[f'A{i}'].font = BOLD if a and not a.startswith('  ') else ARIAL
    ws[f'B{i}'].font = ARIAL
    ws[f'A{i}'].alignment = LEFT
    ws[f'B{i}'].alignment = LEFT

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2: Headcount_Roster (Nov-26 current)
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet('Headcount_Roster')
headers = ['Emp ID','Department','Level','Title','Entity','Start Date','Base Salary','Target Bonus %','Total Target Comp','FTE Status','Band Status','Notes']
widths  = [12,18,8,32,12,13,14,14,16,12,14,30]
for c,(h,w) in enumerate(zip(headers,widths),start=1): hdr(ws,1,c,h,w)

# Roster data: (emp_id, dept, level, title, entity, start_date, base, bonus_pct, status, band_status, notes)
roster = [
    # Engineering — 30 employees (budget = 25, OVER by 5 ← FINDING 1)
    ('EMP-1001','Engineering','L7','CTO','LuminaUS',date(2021,3,15),550000,0.40,'FT','In band',''),
    ('EMP-1002','Engineering','L6','VP Engineering','LuminaUS',date(2022,6,1),420000,0.35,'FT','In band',''),
    ('EMP-1003','Engineering','L5','Director, Platform','LuminaUS',date(2022,9,1),320000,0.25,'FT','In band',''),
    ('EMP-1004','Engineering','L5','Director, Streaming','LuminaUS',date(2023,1,15),310000,0.25,'FT','In band',''),
    ('EMP-1005','Engineering','L5','Director, Data','LuminaUS',date(2023,4,1),315000,0.25,'FT','In band',''),
    ('EMP-1006','Engineering','L4','Eng Manager, Backend','LuminaUS',date(2022,11,1),265000,0.20,'FT','In band',''),
    ('EMP-1007','Engineering','L4','Eng Manager, Frontend','LuminaUS',date(2023,2,1),260000,0.20,'FT','In band',''),
    ('EMP-1008','Engineering','L4','Eng Manager, ML','LuminaUS',date(2023,6,1),270000,0.20,'FT','In band',''),
    ('EMP-1009','Engineering','L4','Eng Manager, DevOps','LuminaUS',date(2024,1,1),255000,0.20,'FT','In band',''),
    ('EMP-1010','Engineering','L4','Eng Manager, Mobile','LuminaUS',date(2024,3,1),258000,0.20,'FT','In band',''),
    ('EMP-1011','Engineering','L3','Sr Software Engineer','LuminaUS',date(2022,5,1),225000,0.15,'FT','In band',''),
    ('EMP-1012','Engineering','L3','Sr Software Engineer','LuminaUS',date(2022,8,1),218000,0.15,'FT','In band',''),
    ('EMP-1013','Engineering','L3','Sr Software Engineer','LuminaUS',date(2022,10,1),222000,0.15,'FT','In band',''),
    ('EMP-1014','Engineering','L3','Sr Software Engineer','LuminaUS',date(2023,1,1),220000,0.15,'FT','In band',''),
    ('EMP-1015','Engineering','L3','Sr Software Engineer','LuminaUS',date(2023,3,1),215000,0.15,'FT','In band',''),
    ('EMP-1016','Engineering','L3','Sr Software Engineer','LuminaUS',date(2023,7,1),219000,0.15,'FT','In band',''),
    ('EMP-1017','Engineering','L3','Sr Software Engineer','LuminaUS',date(2023,9,1),221000,0.15,'FT','In band',''),
    ('EMP-1018','Engineering','L3','Sr Software Engineer','LuminaUS',date(2024,2,1),216000,0.15,'FT','In band',''),
    ('EMP-1019','Engineering','L2','Software Engineer','LuminaUS',date(2023,6,1),178000,0.10,'FT','In band',''),
    ('EMP-1020','Engineering','L2','Software Engineer','LuminaUS',date(2023,8,1),175000,0.10,'FT','In band',''),
    ('EMP-1021','Engineering','L2','Software Engineer','LuminaUS',date(2023,11,1),172000,0.10,'FT','In band',''),
    ('EMP-1022','Engineering','L2','Software Engineer','LuminaUS',date(2024,1,1),180000,0.10,'FT','In band',''),
    ('EMP-1023','Engineering','L2','Software Engineer','LuminaUS',date(2024,3,1),176000,0.10,'FT','In band',''),
    ('EMP-1024','Engineering','L2','Software Engineer','LuminaUS',date(2024,5,1),174000,0.10,'FT','In band',''),
    ('EMP-1025','Engineering','L2','Software Engineer','LuminaUS',date(2024,7,1),177000,0.10,'FT','In band',''),
    # 5 over budget starts here ↓
    ('EMP-1026','Engineering','L2','Software Engineer','LuminaUS',date(2026,8,1),178000,0.10,'FT','In band','OVER BUDGET — headcount above plan'),
    ('EMP-1027','Engineering','L2','Software Engineer','LuminaUS',date(2026,9,1),175000,0.10,'FT','In band','OVER BUDGET — headcount above plan'),
    ('EMP-1028','Engineering','L3','Sr Software Engineer','LuminaUS',date(2026,10,1),220000,0.15,'FT','In band','OVER BUDGET — headcount above plan'),
    ('EMP-1029','Engineering','L2','Software Engineer','LuminaUS',date(2026,10,15),176000,0.10,'FT','In band','OVER BUDGET — headcount above plan'),
    ('EMP-1030','Engineering','L2','Software Engineer','LuminaUS',date(2026,11,1),179000,0.10,'FT','In band','OVER BUDGET — headcount above plan'),
    # G&A — 15 employees (on budget)
    ('EMP-2001','G&A','L7','CFO','LuminaUS',date(2020,7,1),500000,0.50,'FT','In band',''),
    ('EMP-2002','G&A','L7','COO','LuminaUS',date(2021,1,1),480000,0.45,'FT','In band',''),
    ('EMP-2003','G&A','L5','VP Finance','LuminaUS',date(2022,3,1),280000,0.25,'FT','In band',''),
    ('EMP-2004','G&A','L5','VP People & HR','LuminaUS',date(2022,5,1),275000,0.25,'FT','In band',''),
    ('EMP-2005','G&A','L4','Director, Accounting','LuminaUS',date(2023,2,1),205000,0.20,'FT','In band',''),
    ('EMP-2006','G&A','L4','Director, FP&A','LuminaUS',date(2023,4,1),210000,0.20,'FT','In band',''),
    ('EMP-2007','G&A','L3','Sr Finance Manager','LuminaUS',date(2023,6,1),165000,0.15,'FT','In band',''),
    ('EMP-2008','G&A','L3','Sr HR Manager','LuminaUS',date(2023,8,1),160000,0.15,'FT','In band',''),
    ('EMP-2009','G&A','L3','Sr Legal Counsel','LuminaUS',date(2024,1,1),170000,0.15,'FT','In band',''),
    ('EMP-2010','G&A','L3','Controller','LuminaUS',date(2024,2,1),168000,0.15,'FT','In band',''),
    ('EMP-2011','G&A','L2','Finance Analyst','LuminaUS',date(2024,6,1),88000,0.10,'FT','Below band','FINDING 4 — paid 18% below midpoint ($88K vs $107K mid)'),
    ('EMP-2012','G&A','L2','HR Generalist','LuminaUS',date(2024,7,1),95000,0.10,'FT','In band',''),
    ('EMP-2013','G&A','L2','Accounting Analyst','LuminaUS',date(2024,9,1),97000,0.10,'FT','In band',''),
    ('EMP-2014','G&A','L2','Legal Analyst','LuminaUS',date(2025,1,1),100000,0.10,'FT','In band',''),
    ('EMP-2015','G&A','L2','FP&A Analyst','LuminaUS',date(2025,3,1),105000,0.10,'FT','In band',''),
    # Sales — 14 employees (1 open, see pipeline)
    ('EMP-3001','Sales','L7','CRO','LuminaUS',date(2021,6,1),450000,0.60,'FT','In band',''),
    ('EMP-3002','Sales','L5','VP Sales, Enterprise','LuminaUS',date(2022,8,1),265000,0.40,'FT','In band',''),
    ('EMP-3003','Sales','L5','VP Sales, SMB','LuminaUS',date(2023,1,1),258000,0.40,'FT','In band',''),
    ('EMP-3004','Sales','L4','Sales Manager, East','LuminaUS',date(2023,3,1),155000,0.30,'FT','In band',''),
    ('EMP-3005','Sales','L4','Sales Manager, West','LuminaUS',date(2023,5,1),152000,0.30,'FT','In band',''),
    ('EMP-3006','Sales','L4','Sales Manager, Central','LuminaUS',date(2023,9,1),150000,0.30,'FT','In band',''),
    ('EMP-3007','Sales','L3','Sr Account Executive','LuminaUS',date(2023,2,1),105000,0.50,'FT','In band',''),
    ('EMP-3008','Sales','L3','Sr Account Executive','LuminaUS',date(2023,6,1),102000,0.50,'FT','In band',''),
    ('EMP-3009','Sales','L3','Sr Account Executive','LuminaUS',date(2024,1,1),104000,0.50,'FT','In band',''),
    ('EMP-3010','Sales','L3','Sr Account Executive','LuminaUS',date(2024,4,1),103000,0.50,'FT','In band',''),
    ('EMP-3011','Sales','L2','Account Executive','LuminaUS',date(2024,3,1),82000,0.60,'FT','In band',''),
    ('EMP-3012','Sales','L2','Account Executive','LuminaUS',date(2024,6,1),80000,0.60,'FT','In band',''),
    ('EMP-3013','Sales','L2','Account Executive','LuminaUS',date(2024,9,1),83000,0.60,'FT','In band',''),
    ('EMP-3014','Sales','L2','Sales Development Rep','LuminaUS',date(2025,1,1),65000,0.40,'FT','In band',''),
    # Marketing — 10 employees (2 out-of-band high ← FINDING 3)
    ('EMP-4001','Marketing','L7','CMO','LuminaUS',date(2021,9,1),470000,0.45,'FT','In band',''),
    ('EMP-4002','Marketing','L5','VP Brand Marketing','LuminaUS',date(2022,11,1),255000,0.25,'FT','In band',''),
    ('EMP-4003','Marketing','L5','VP Performance Mktg','LuminaUS',date(2023,2,1),250000,0.25,'FT','In band',''),
    ('EMP-4004','Marketing','L4','Director, Brand','LuminaUS',date(2023,5,1),190000,0.20,'FT','In band',''),
    ('EMP-4005','Marketing','L3','Sr Marketing Manager','LuminaUS',date(2023,7,1),148000,0.15,'FT','In band',''),
    ('EMP-4006','Marketing','L3','Sr Marketing Manager','LuminaUS',date(2023,10,1),152000,0.15,'FT','In band',''),
    # Out-of-band high ↓
    ('EMP-4007','Marketing','L3','Marketing Manager','LuminaUS',date(2024,2,1),168000,0.15,'FT','Above band','FINDING 3 — paid $168K vs band midpoint $138K (+22%)'),
    ('EMP-4008','Marketing','L3','Marketing Manager','LuminaUS',date(2024,5,1),162000,0.15,'FT','Above band','FINDING 3 — paid $162K vs band midpoint $138K (+17%)'),
    ('EMP-4009','Marketing','L2','Marketing Specialist','LuminaUS',date(2024,8,1),95000,0.10,'FT','In band',''),
    ('EMP-4010','Marketing','L2','Content Strategist','LuminaUS',date(2025,2,1),92000,0.10,'FT','In band',''),
    # Content — 11 employees (3 departed in Nov, currently 8 active ← FINDING 5)
    ('EMP-5001','Content','L6','SVP Content','LuminaUS',date(2021,4,1),410000,0.35,'FT','In band',''),
    ('EMP-5002','Content','L5','VP Original Content','LuminaUS',date(2022,7,1),285000,0.30,'FT','In band',''),
    ('EMP-5003','Content','L5','VP Content Strategy','LuminaUS',date(2023,3,1),278000,0.30,'FT','In band',''),
    ('EMP-5004','Content','L4','Director, Originals','LuminaUS',date(2023,8,1),185000,0.20,'FT','In band',''),
    ('EMP-5005','Content','L4','Director, Licensing','LuminaUS',date(2024,1,1),180000,0.20,'FT','In band',''),
    ('EMP-5006','Content','L3','Sr Content Manager','LuminaUS',date(2024,3,1),150000,0.15,'FT','In band',''),
    ('EMP-5007','Content','L3','Sr Content Manager','LuminaUS',date(2024,6,1),148000,0.15,'FT','In band',''),
    ('EMP-5008','Content','L2','Content Associate','LuminaUS',date(2025,1,1),88000,0.10,'FT','In band',''),
]

for r, row in enumerate(roster, start=2):
    emp, dept, lvl, title, entity, start, base, bonus, fte, band_status, notes = row
    cell(ws, r, 1, emp)
    cell(ws, r, 2, dept)
    cell(ws, r, 3, lvl, bold=False)
    cell(ws, r, 4, title)
    cell(ws, r, 5, entity)
    cell(ws, r, 6, start, DATE_FMT)
    cell(ws, r, 7, base, CURRENCY)
    cell(ws, r, 8, bonus, PCT)
    ws.cell(row=r, column=9, value=f'=G{r}*(1+H{r})').number_format = CURRENCY
    ws.cell(row=r, column=9).border = BOX
    ws.cell(row=r, column=9).font = ARIAL
    cell(ws, r, 10, fte)
    c11 = cell(ws, r, 11, band_status)
    cell(ws, r, 12, notes, wrap=True)
    if 'OVER BUDGET' in notes:
        for c in range(1,13): ws.cell(row=r,column=c).fill = WARN_FILL
    elif 'FINDING 3' in notes:
        for c in range(1,13): ws.cell(row=r,column=c).fill = FLAG_FILL
    elif 'FINDING 4' in notes:
        for c in range(1,13): ws.cell(row=r,column=c).fill = FLAG_FILL

ws.freeze_panes = 'A2'

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 3: Prior_Roster (Oct-26 — includes 3 Nov departures)
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet('Prior_Roster')
for c,(h,w) in enumerate(zip(headers,widths),start=1): hdr(ws,1,c,h,w)

# Same as current roster but with 3 additional Content employees who departed in Nov
prior_additions = [
    ('EMP-5009','Content','L3','Sr Content Manager','LuminaUS',date(2023,9,1),152000,0.15,'FT','In band','DEPARTED Nov 2026 — voluntary'),
    ('EMP-5010','Content','L2','Content Associate','LuminaUS',date(2024,11,1),86000,0.10,'FT','In band','DEPARTED Nov 2026 — voluntary'),
    ('EMP-5011','Content','L2','Content Coordinator','LuminaUS',date(2025,2,1),82000,0.10,'FT','In band','DEPARTED Nov 2026 — voluntary'),
]
prior_roster = roster + prior_additions
for r, row in enumerate(prior_roster, start=2):
    emp, dept, lvl, title, entity, start, base, bonus, fte, band_status, notes = row
    cell(ws, r, 1, emp); cell(ws, r, 2, dept); cell(ws, r, 3, lvl)
    cell(ws, r, 4, title); cell(ws, r, 5, entity); cell(ws, r, 6, start, DATE_FMT)
    cell(ws, r, 7, base, CURRENCY); cell(ws, r, 8, bonus, PCT)
    ws.cell(row=r, column=9, value=f'=G{r}*(1+H{r})').number_format = CURRENCY
    ws.cell(row=r, column=9).border = BOX; ws.cell(row=r, column=9).font = ARIAL
    cell(ws, r, 10, fte); cell(ws, r, 11, band_status); cell(ws, r, 12, notes, wrap=True)
    if 'DEPARTED' in notes:
        for c in range(1,13): ws.cell(row=r,column=c).fill = WARN_FILL

ws.freeze_panes = 'A2'

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 4: Headcount_Budget
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet('Headcount_Budget')
months = ['Jan-26','Feb-26','Mar-26','Apr-26','May-26','Jun-26','Jul-26','Aug-26','Sep-26','Oct-26','Nov-26','Dec-26']
bud_headers = ['Department','Metric'] + months + ['FY2026 Total']
bud_widths  = [18,16] + [10]*12 + [14]
for c,(h,w) in enumerate(zip(bud_headers,bud_widths),start=1): hdr(ws,1,c,h,w)

# Budget data: dept, HC budget, HC actual, cost budget, cost actual by month
# Nov-26 actuals: Engineering 30 (budget 25), Content 8 (budget 11 after attritions)
budget_data = [
    # Engineering
    ('Engineering','HC Budget',    25,25,25,25,25,25,25,25,25,25,25,25),
    ('Engineering','HC Actual',    25,25,25,25,25,25,26,27,28,29,30,30),
    ('Engineering','Cost Budget ($)', 1400000,1400000,1400000,1400000,1400000,1400000,1400000,1400000,1400000,1400000,1400000,1400000),
    ('Engineering','Cost Actual ($)', 1400000,1400000,1400000,1400000,1400000,1400000,1440000,1480000,1530000,1570000,1620000,1620000),
    # G&A
    ('G&A','HC Budget',            15,15,15,15,15,15,15,15,15,15,15,15),
    ('G&A','HC Actual',            15,15,15,15,15,15,15,15,15,15,15,15),
    ('G&A','Cost Budget ($)',       750000,750000,750000,750000,750000,750000,750000,750000,750000,750000,750000,750000),
    ('G&A','Cost Actual ($)',       750000,750000,750000,750000,750000,750000,750000,750000,750000,750000,750000,750000),
    # Sales
    ('Sales','HC Budget',          15,15,15,15,15,15,15,15,15,15,15,15),
    ('Sales','HC Actual',          15,15,15,15,15,15,14,14,14,14,14,14),
    ('Sales','Cost Budget ($)',     850000,850000,850000,850000,850000,850000,850000,850000,850000,850000,850000,850000),
    ('Sales','Cost Actual ($)',     850000,850000,850000,850000,850000,850000,800000,800000,800000,800000,800000,800000),
    # Marketing
    ('Marketing','HC Budget',      10,10,10,10,10,10,10,10,10,10,10,10),
    ('Marketing','HC Actual',      10,10,10,10,10,10,10,10,10,10,10,10),
    ('Marketing','Cost Budget ($)', 600000,600000,600000,600000,600000,600000,600000,600000,600000,600000,600000,600000),
    ('Marketing','Cost Actual ($)', 600000,600000,600000,600000,600000,600000,600000,600000,600000,600000,625000,625000),
    # Content
    ('Content','HC Budget',        11,11,11,11,11,11,11,11,11,11,11,11),
    ('Content','HC Actual',        11,11,11,11,11,11,11,11,11,11,8,8),
    ('Content','Cost Budget ($)',   650000,650000,650000,650000,650000,650000,650000,650000,650000,650000,650000,650000),
    ('Content','Cost Actual ($)',   650000,650000,650000,650000,650000,650000,650000,650000,650000,650000,480000,480000),
]

for r, row in enumerate(budget_data, start=2):
    dept, metric = row[0], row[1]
    cell(ws, r, 1, dept, bold=True)
    cell(ws, r, 2, metric)
    for i, val in enumerate(row[2:], start=3):
        fmt = CURRENCY if 'Cost' in metric else INT_FMT
        cell(ws, r, i, val, fmt)
        # Flag Nov-26 Engineering actual over budget
        if i == 14 and dept == 'Engineering' and metric == 'HC Actual':
            ws.cell(row=r, column=i).fill = FLAG_FILL
        # Flag Nov-26 Content actual under budget (attrition)
        if i == 14 and dept == 'Content' and metric == 'HC Actual':
            ws.cell(row=r, column=i).fill = WARN_FILL
    # FY total
    col_ltr_start = get_column_letter(3)
    col_ltr_end = get_column_letter(14)
    fmt = CURRENCY if 'Cost' in metric else INT_FMT
    f = ws.cell(row=r, column=15, value=f'=SUM(C{r}:N{r})')
    f.number_format = fmt; f.border = BOX; f.font = BOLD; f.fill = TOTAL_FILL

ws.freeze_panes = 'C2'

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 5: Open_Pipeline
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet('Open_Pipeline')
pipe_headers = ['Req ID','Department','Level','Title','Entity','Req Open Date','Target Start','Days Past Target','Budgeted Base','Status','Hiring Manager','Notes']
pipe_widths  = [12,14,8,30,10,13,13,16,14,16,22,35]
for c,(h,w) in enumerate(zip(pipe_headers,pipe_widths),start=1): hdr(ws,1,c,h,w)

pipeline = [
    # Engineering — 3 flagged past target start ← FINDING 2
    ('REQ-101','Engineering','L2','Software Engineer','LuminaUS',date(2026,7,1),date(2026,9,1),90,178000,'Interviewing','EMP-1006','FINDING 2 — 90 days past target start'),
    ('REQ-102','Engineering','L3','Sr Software Engineer','LuminaUS',date(2026,8,1),date(2026,10,1),61,220000,'Offer Extended','EMP-1007','FINDING 2 — 61 days past target start'),
    ('REQ-103','Engineering','L2','Software Engineer','LuminaUS',date(2026,9,1),date(2026,10,15),46,176000,'Sourcing','EMP-1008','FINDING 2 — 46 days past target start'),
    # Engineering — on track
    ('REQ-104','Engineering','L4','Eng Manager, Security','LuminaUS',date(2026,10,15),date(2026,12,1),0,265000,'Final Interviews','EMP-1002','On track'),
    ('REQ-105','Engineering','L3','Sr Software Engineer','LuminaUS',date(2026,11,1),date(2027,1,15),0,222000,'Sourcing','EMP-1003','On track'),
    # Sales — 1 flagged ← FINDING 6
    ('REQ-201','Sales','L2','Account Executive','LuminaUS',date(2026,9,15),date(2026,10,15),46,82000,'Offer Stage','EMP-3002','FINDING 6 — 45 days past target, commission plan not assigned'),
    ('REQ-202','Sales','L3','Sr Account Executive','LuminaUS',date(2026,11,1),date(2027,1,1),0,104000,'Sourcing','EMP-3003','On track'),
    # G&A
    ('REQ-301','G&A','L2','FP&A Analyst','LuminaUS',date(2026,10,1),date(2026,12,1),0,102000,'Interviewing','EMP-2006','On track'),
    ('REQ-302','G&A','L3','Sr HR Manager','LuminaUS',date(2026,11,15),date(2027,2,1),0,162000,'Sourcing','EMP-2004','On track'),
    # Content — backfilling Nov departures
    ('REQ-501','Content','L3','Sr Content Manager','LuminaUS',date(2026,11,15),date(2027,1,15),0,150000,'Sourcing','EMP-5002','Backfill for EMP-5009 departure'),
    ('REQ-502','Content','L2','Content Associate','LuminaUS',date(2026,11,20),date(2027,2,1),0,88000,'Not Yet Open','EMP-5002','Backfill pending HR approval'),
    # Marketing
    ('REQ-401','Marketing','L2','Marketing Specialist','LuminaUS',date(2026,11,1),date(2027,1,15),0,95000,'Sourcing','EMP-4003','New headcount — approved in Q4 plan'),
    # EMEA
    ('REQ-601','Engineering','L3','Sr Software Engineer','LuminaEMEA',date(2026,10,1),date(2026,12,1),0,195000,'Interviewing','EMP-1002','EMEA hire — EUR comp converted to USD'),
    # APAC
    ('REQ-701','Engineering','L2','Software Engineer','LuminaAPAC',date(2026,11,1),date(2027,1,15),0,145000,'Sourcing','EMP-1002','APAC hire — SGD comp converted to USD'),
    ('REQ-702','Sales','L2','Account Executive','LuminaAPAC',date(2026,11,15),date(2027,2,1),0,72000,'Sourcing','EMP-3003','APAC hire'),
]

for r, row in enumerate(pipeline, start=2):
    req, dept, lvl, title, entity, open_dt, target_dt, days_past, budget_base, status, mgr, notes = row
    cell(ws, r, 1, req); cell(ws, r, 2, dept); cell(ws, r, 3, lvl)
    cell(ws, r, 4, title); cell(ws, r, 5, entity)
    cell(ws, r, 6, open_dt, DATE_FMT); cell(ws, r, 7, target_dt, DATE_FMT)
    cell(ws, r, 8, days_past, INT_FMT)
    cell(ws, r, 9, budget_base, CURRENCY)
    cell(ws, r, 10, status); cell(ws, r, 11, mgr)
    cell(ws, r, 12, notes, wrap=True)
    if days_past > 30:
        for c in range(1,13): ws.cell(row=r,column=c).fill = FLAG_FILL
    elif 'FINDING 6' in notes:
        for c in range(1,13): ws.cell(row=r,column=c).fill = WARN_FILL

ws.freeze_panes = 'A2'

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 6: Comp_Bands
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet('Comp_Bands')
band_headers = ['Department','Level','Role Category','Band Min','Band Midpoint','Band Max','Spread %']
band_widths  = [18,8,24,14,14,14,11]
for c,(h,w) in enumerate(zip(band_headers,band_widths),start=1): hdr(ws,1,c,h,w)

bands = [
    ('Engineering','L2','Software Engineer',        155000,175000,195000),
    ('Engineering','L3','Sr Software Engineer',     195000,215000,240000),
    ('Engineering','L4','Engineering Manager',      240000,260000,285000),
    ('Engineering','L5','Director, Engineering',    290000,315000,345000),
    ('Engineering','L6','VP Engineering',           380000,420000,470000),
    ('Engineering','L7','CTO',                      500000,550000,620000),
    ('G&A','L2','Finance / HR Analyst',             90000,107000,125000),
    ('G&A','L3','Sr Finance / HR Manager',          145000,162000,180000),
    ('G&A','L4','Director, Finance / HR',           185000,205000,230000),
    ('G&A','L5','VP Finance / People',              255000,278000,310000),
    ('G&A','L7','CFO / COO',                        440000,490000,560000),
    ('Sales','L2','Account Executive',              72000,82000,95000),
    ('Sales','L3','Sr Account Executive',           92000,104000,118000),
    ('Sales','L4','Sales Manager',                  138000,153000,170000),
    ('Sales','L5','VP Sales',                       240000,260000,290000),
    ('Sales','L7','CRO',                            400000,450000,510000),
    ('Marketing','L2','Marketing Specialist',       82000,95000,110000),
    ('Marketing','L3','Marketing Manager',          118000,138000,158000),
    ('Marketing','L4','Director, Marketing',        168000,188000,212000),
    ('Marketing','L5','VP Marketing',               230000,252000,280000),
    ('Marketing','L7','CMO',                        420000,468000,530000),
    ('Content','L2','Content Associate',            78000,90000,105000),
    ('Content','L3','Sr Content Manager',           130000,148000,168000),
    ('Content','L4','Director, Content',            165000,183000,205000),
    ('Content','L5','VP Content',                   255000,278000,310000),
    ('Content','L6','SVP Content',                  365000,405000,455000),
]

for r, row in enumerate(bands, start=2):
    dept, lvl, role, bmin, bmid, bmax = row
    cell(ws, r, 1, dept); cell(ws, r, 2, lvl); cell(ws, r, 3, role)
    cell(ws, r, 4, bmin, CURRENCY); cell(ws, r, 5, bmid, CURRENCY); cell(ws, r, 6, bmax, CURRENCY)
    f = ws.cell(row=r, column=7, value=f'=(F{r}-D{r})/D{r}')
    f.number_format = PCT; f.border = BOX; f.font = ARIAL

ws.freeze_panes = 'A2'

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 7: Market_Benchmarks
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet('Market_Benchmarks')
mkt_headers = ['Role Category','Level','Market P25','Market P50 (Median)','Market P75','Source','Last Updated']
mkt_widths  = [28,8,14,16,14,18,13]
for c,(h,w) in enumerate(zip(mkt_headers,mkt_widths),start=1): hdr(ws,1,c,h,w)

benchmarks = [
    ('Software Engineer','L2',148000,168000,192000,'Indeed MCP / Levels.fyi',date(2026,10,1)),
    ('Sr Software Engineer','L3',188000,210000,238000,'Indeed MCP / Levels.fyi',date(2026,10,1)),
    ('Engineering Manager','L4',232000,255000,282000,'Indeed MCP / Radford',date(2026,10,1)),
    ('Director, Engineering','L5',280000,308000,342000,'Indeed MCP / Radford',date(2026,10,1)),
    ('VP Engineering','L6',368000,412000,462000,'Indeed MCP / Radford',date(2026,10,1)),
    ('Finance / HR Analyst','L2',85000,98000,116000,'Indeed MCP / Radford',date(2026,10,1)),
    ('Sr Finance / HR Manager','L3',140000,157000,178000,'Indeed MCP / Radford',date(2026,10,1)),
    ('Director, Finance / HR','L4',178000,198000,224000,'Indeed MCP / Radford',date(2026,10,1)),
    ('VP Finance / People','L5',248000,272000,305000,'Indeed MCP / Radford',date(2026,10,1)),
    ('Account Executive','L2',68000,78000,92000,'Indeed MCP',date(2026,10,1)),
    ('Sr Account Executive','L3',88000,100000,115000,'Indeed MCP',date(2026,10,1)),
    ('Sales Manager','L4',132000,148000,166000,'Indeed MCP / Radford',date(2026,10,1)),
    ('VP Sales','L5',232000,255000,285000,'Indeed MCP / Radford',date(2026,10,1)),
    ('Marketing Specialist','L2',78000,91000,108000,'Indeed MCP',date(2026,10,1)),
    ('Marketing Manager','L3',112000,132000,155000,'Indeed MCP',date(2026,10,1)),
    ('Director, Marketing','L4',162000,182000,208000,'Indeed MCP',date(2026,10,1)),
    ('VP Marketing','L5',222000,246000,275000,'Indeed MCP',date(2026,10,1)),
    ('Content Associate','L2',74000,86000,102000,'Indeed MCP',date(2026,10,1)),
    ('Sr Content Manager','L3',124000,142000,164000,'Indeed MCP',date(2026,10,1)),
    ('Director, Content','L4',158000,176000,200000,'Indeed MCP',date(2026,10,1)),
    ('VP Content','L5',242000,268000,302000,'Indeed MCP',date(2026,10,1)),
]

for r, row in enumerate(benchmarks, start=2):
    role, lvl, p25, p50, p75, source, updated = row
    cell(ws, r, 1, role); cell(ws, r, 2, lvl)
    cell(ws, r, 3, p25, CURRENCY); cell(ws, r, 4, p50, CURRENCY); cell(ws, r, 5, p75, CURRENCY)
    cell(ws, r, 6, source); cell(ws, r, 7, updated, DATE_FMT)

ws.freeze_panes = 'A2'

# ─────────────────────────────────────────────────────────────────────────────
# Sheet 8: Attrition_Log
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet('Attrition_Log')
att_headers = ['Emp ID','Department','Level','Title','Entity','Last Day','Departure Type','Reason (Category)','Regrettable?','Notes']
att_widths  = [12,14,8,28,10,12,16,22,12,38]
for c,(h,w) in enumerate(zip(att_headers,att_widths),start=1): hdr(ws,1,c,h,w)

attrition = [
    # November — Content team 3 departures ← FINDING 5
    ('EMP-5009','Content','L3','Sr Content Manager','LuminaUS',date(2026,11,7),'Voluntary','New opportunity — competitor',True,'FINDING 5 — 1 of 3 Nov Content departures. Annualised rate 36%'),
    ('EMP-5010','Content','L2','Content Associate','LuminaUS',date(2026,11,14),'Voluntary','Personal reasons',False,'FINDING 5 — 2 of 3 Nov Content departures'),
    ('EMP-5011','Content','L2','Content Coordinator','LuminaUS',date(2026,11,21),'Voluntary','Returning to school',False,'FINDING 5 — 3 of 3 Nov Content departures. Trigger: >2% monthly'),
    # October
    ('EMP-1031','Engineering','L2','Software Engineer','LuminaUS',date(2026,10,3),'Voluntary','Compensation — competing offer',True,'Left for 20% base increase. Comp band review flagged.'),
    # September
    ('EMP-3015','Sales','L2','Account Executive','LuminaUS',date(2026,9,15),'Involuntary','Performance',False,'PIP outcome. Headcount not backfilled — open REQ-201'),
    ('EMP-2016','G&A','L2','Finance Analyst','LuminaUS',date(2026,9,30),'Voluntary','New opportunity',True,'Lateral move. Vacancy replaced by EMP-2011'),
]

for r, row in enumerate(attrition, start=2):
    emp, dept, lvl, title, entity, last_day, dep_type, reason, regrettable, notes = row
    cell(ws, r, 1, emp); cell(ws, r, 2, dept); cell(ws, r, 3, lvl)
    cell(ws, r, 4, title); cell(ws, r, 5, entity)
    cell(ws, r, 6, last_day, DATE_FMT)
    cell(ws, r, 7, dep_type); cell(ws, r, 8, reason)
    cell(ws, r, 9, 'Yes' if regrettable else 'No')
    cell(ws, r, 10, notes, wrap=True)
    if 'FINDING 5' in notes:
        for c in range(1,11): ws.cell(row=r,column=c).fill = FLAG_FILL
    elif 'Compensation' in reason:
        for c in range(1,11): ws.cell(row=r,column=c).fill = WARN_FILL

ws.freeze_panes = 'A2'

# ─────────────────────────────────────────────────────────────────────────────
# Save
# ─────────────────────────────────────────────────────────────────────────────
out = '/home/claude/lumina_headcount_dataset.xlsx'
wb.save(out)
print(f'Saved: {out}')
